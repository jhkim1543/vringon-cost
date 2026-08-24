# -*- coding: utf-8 -*-
"""워크북(신발_BOM_Cost_소재DB_최종_v2)을 서버가 읽는 JSON 시드로 변환한다.

    python tools/seed_from_xlsx.py [xlsx경로]

워크북은 시트마다 제목/설명 줄이 앞에 붙어 있고 헤더 행 위치가 다르다.
헤더는 "비어 있지 않은 셀이 가장 많은 초기 행"으로 자동 탐지한다.
값은 원본 그대로 싣고, 단위·통화 정규화는 서버 쪽 로더가 담당한다.
"""
import json
import sys
import datetime as dt
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path.home() / "Downloads" / "신발_BOM_Cost_소재DB_최종_v2_2026-08-23.xlsx"

# 시트명 -> 출력 파일명. 워크북에 없는 시트는 조용히 건너뛴다.
SHEETS = {
    "01_파트소재맵": "part_material_map.json",
    "02_단가관측": "price_observations.json",
    "03_원자재지수": "raw_material_index.json",
    "06_FX단위환산": "fx.json",
    "10_신발BOM마스터": "bom_master.json",
    "14_ConstructionRecipe": "construction_recipes.json",
    "15_공정인건비": "routing.json",
    "16_분기기준단가": "quarterly_prices.json",
    "17_Tooling마스터": "tooling.json",
}


def _cell(v):
    """엑셀 값을 JSON 안전한 값으로."""
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v


def _find_header(ws, scan=12):
    """앞쪽 scan행 중 채워진 셀이 가장 많은 행을 헤더로 본다."""
    best, best_n = 1, -1
    for r in range(1, min(scan, ws.max_row) + 1):
        row = [_cell(c) for c in next(ws.iter_rows(min_row=r, max_row=r, values_only=True))]
        n = sum(1 for c in row if c is not None)
        # 제목 줄은 보통 1칸만 차지한다. 동점이면 위쪽 행을 택한다.
        if n > best_n:
            best, best_n = r, n
    return best


def extract(ws):
    hdr_row = _find_header(ws)
    headers = [_cell(c) for c in next(ws.iter_rows(min_row=hdr_row, max_row=hdr_row, values_only=True))]
    # 뒤쪽 빈 컬럼 잘라내기
    while headers and headers[-1] is None:
        headers.pop()
    if not headers:
        return [], hdr_row

    rows = []
    for raw in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        vals = [_cell(c) for c in raw[: len(headers)]]
        if all(v is None for v in vals):
            continue
        rec = {}
        for h, v in zip(headers, vals):
            if h is None:
                continue
            rec[h] = v
        # 첫 컬럼이 비어 있으면 표 하단의 주석/합계 줄로 보고 버린다.
        if headers[0] is not None and rec.get(headers[0]) is None:
            continue
        rows.append(rec)
    return rows, hdr_row


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.exists():
        sys.exit(f"워크북을 찾을 수 없습니다: {xlsx}")

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    out_dir = ROOT / "data" / "seed"
    out_dir.mkdir(parents=True, exist_ok=True)

    manifest = {
        "source_workbook": xlsx.name,
        "extracted_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "sheets": {},
    }

    for sheet, fname in SHEETS.items():
        if sheet not in wb.sheetnames:
            print(f"  skip  {sheet} (없음)")
            continue
        rows, hdr = extract(wb[sheet])
        (out_dir / fname).write_text(
            json.dumps(rows, ensure_ascii=False, indent=1), encoding="utf-8"
        )
        manifest["sheets"][sheet] = {"file": fname, "header_row": hdr, "rows": len(rows)}
        print(f"  ok    {sheet:24s} header@{hdr}  {len(rows):3d}행 -> data/seed/{fname}")

    (out_dir / "_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=1), encoding="utf-8"
    )
    print(f"\n총 {len(manifest['sheets'])}개 시트 추출 완료 -> {out_dir}")


if __name__ == "__main__":
    main()
